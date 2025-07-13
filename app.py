from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import pandas as pd
import os, tempfile, re, datetime
from google.cloud import documentai_v1 as documentai
import openpyxl
from openpyxl import load_workbook
from copy import copy

# === CONFIG (cloud-ready) ===
PROJECT_ID = os.environ.get('PROJECT_ID', 'driver-schedule-ocr')
LOCATION = os.environ.get('LOCATION', 'us')
PROCESSOR_ID = os.environ.get('PROCESSOR_ID', '2acb7269aa33ccf9')

# Path to Google Service Account key file (set as ENV on cloud)
GOOGLE_KEY_PATH = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "/etc/secrets/driver-schedule-ocr.json")
os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = GOOGLE_KEY_PATH

# Path to Excel template (set as ENV or place in app root)
TEMPLATE_PATH = os.environ.get("TEMPLATE_PATH", "Truck_Load_Record_Template.xlsx")
# ============================

app = Flask(__name__)
CORS(app)

@app.route("/")
def home():
    return "Hello, Flask! The server is running."

def extract_table_from_image(image_path):
    client = documentai.DocumentProcessorServiceClient()
    name = f"projects/{PROJECT_ID}/locations/{LOCATION}/processors/{PROCESSOR_ID}"
    with open(image_path, "rb") as image:
        image_content = image.read()
    raw_document = documentai.RawDocument(content=image_content, mime_type="image/jpeg")
    request = documentai.ProcessRequest(name=name, raw_document=raw_document)
    result = client.process_document(request=request)
    document = result.document

    # Extract tables
    tables = []
    headers = []
    for page in document.pages:
        for table in page.tables:
            header_cells = []
            for cell in table.header_rows[0].cells:
                cell_text = ""
                for segment in cell.layout.text_anchor.text_segments:
                    start = int(segment.start_index)
                    end = int(segment.end_index)
                    cell_text += document.text[start:end]
                header_cells.append(cell_text.strip())
            headers = header_cells
            for row in table.body_rows:
                row_values = []
                for cell in row.cells:
                    cell_text = ""
                    for segment in cell.layout.text_anchor.text_segments:
                        start = int(segment.start_index)
                        end = int(segment.end_index)
                        cell_text += document.text[start:end]
                    row_values.append(cell_text.strip())
                tables.append(row_values)
    if not tables:
        return None
    df = pd.DataFrame(tables, columns=headers)

    # Robust column matching
    def find_col(df, substrings):
        for target in substrings:
            for col in df.columns:
                clean_col = col.lower().replace(" ", "").replace("_", "")
                if target in clean_col:
                    return col
        return None

    # Try multiple variants for each required field
    run_col     = find_col(df, ["run", "run#", "runno", "runno.", "runnumber"])
    driver1_col = find_col(df, ["driver1", "driver 1", "driver", "drivers", "drivername"])
    driver2_col = find_col(df, ["driver2", "driver 2", "codriver", "co-driver"])
    truck_col   = find_col(df, ["truck", "vehicle", "reg", "rego", "regno", "registration"])

    # For logging and debugging, print or log extracted columns
    print("Extracted columns:", list(df.columns))

    # If required columns are missing, return a clear error
    missing_cols = []
    if not run_col:     missing_cols.append("Run Number")
    if not driver1_col: missing_cols.append("Driver 1")
    if not truck_col:   missing_cols.append("Truck/Rego")
    if missing_cols:
        raise Exception(
            f"Missing required columns: {', '.join(missing_cols)}. "
            f"Extracted columns: {list(df.columns)}"
        )

    # Cleaners (same as before)
    def extract_run_numbers(cell):
        nums = re.findall(r'\b\d{4}\b', str(cell))
        return " / ".join(nums)

    def clean_driver(cell):
        if pd.isnull(cell): return ""
        name = str(cell).split('\n')[0]
        return re.sub(r'[^A-Za-z ]+', '', name).strip()

    def clean_truck(cell):
        cell = str(cell).strip()
        return cell[:8]  # Slightly longer for full regos

    df_clean = pd.DataFrame()
    df_clean["Run#"]    = df[run_col].apply(extract_run_numbers)
    df_clean["Driver1"] = df[driver1_col].apply(clean_driver)
    if driver2_col:
        df_clean["Driver2"] = df[driver2_col].apply(clean_driver)
    else:
        df_clean["Driver2"] = ""
    df_clean["Truck"]   = df[truck_col].apply(clean_truck)

    return df_clean

def fill_template_per_truck(df_clean):
    from openpyxl import load_workbook
    import tempfile
    import datetime

    template_wb = load_workbook(TEMPLATE_PATH)
    template_sheet = template_wb.active

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)
    today = datetime.date.today() + datetime.timedelta(days=1)

    for truck_row in df_clean.to_dict(orient="records"):
        sheet_name = (truck_row.get("Truck") or truck_row.get("Run#") or "Sheet")[:31]
        ws = out_wb.create_sheet(title=sheet_name)

        # Copy cell values only, no formatting!
        for row in template_sheet.iter_rows():
            for cell in row:
                ws[cell.coordinate].value = cell.value

        # Copy merged cells
        for merged_cell in template_sheet.merged_cells.ranges:
            ws.merge_cells(str(merged_cell))

        # Copy column widths (optional, safe)
        for col_letter, dim in template_sheet.column_dimensions.items():
            ws.column_dimensions[col_letter].width = dim.width

        # Copy row heights (optional, safe)
        for row_idx, dim in template_sheet.row_dimensions.items():
            ws.row_dimensions[row_idx].height = dim.height

        # Fill YOUR fields
        ws["B3"] = truck_row.get("Run#", "")
        ws["I3"] = truck_row.get("Truck", "")
        driver1 = truck_row.get("Driver1", "")
        driver2 = truck_row.get("Driver2", "")
        ws["B4"] = " / ".join([d for d in [driver1, driver2] if d])
        ws["I4"] = today.strftime("%d/%m/%Y")

    out = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    out_wb.save(out.name)
    out.close()
    return out.name


@app.route('/parse_schedule_excel', methods=['POST'])
def parse_schedule_excel():
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
    file = request.files['file']
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as temp:
            file.save(temp.name)
            df_clean = extract_table_from_image(temp.name)
        if df_clean is None or df_clean.empty:
            return jsonify({"error": "No table detected"}), 400

        filled_path = fill_template_per_truck(df_clean)
        return send_file(filled_path, as_attachment=True, download_name="truck_load_records.xlsx")
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))  # Use Render's $PORT, or 5000 locally
    app.run(host="0.0.0.0", port=port)
